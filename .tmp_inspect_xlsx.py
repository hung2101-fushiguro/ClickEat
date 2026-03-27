from openpyxl import load_workbook
p = r"report/Template1_ProjectTracking (2).xlsx"
wb = load_workbook(p)
print('Sheets:', wb.sheetnames)
for ws in wb.worksheets:
    print(f"\n== {ws.title} ==")
    print('max_row', ws.max_row, 'max_col', ws.max_column)
    for r in range(1, min(ws.max_row, 80) + 1):
        row = [ws.cell(r,c).value for c in range(1, min(ws.max_column, 20)+1)]
        if any(v not in (None, '') for v in row):
            print(r, row)
